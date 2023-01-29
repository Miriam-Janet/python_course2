import openpyxl
import random

#Manda a traer la lista de los nombres del txt
list_name = open('names.txt','r')
lista = list_name.read() # Lee todo el archivo
name = lista.split("\n")#el texto lo convierte en lista, aqui se convertira por los saltos de linea
lename = len(name)#cuenta la cantidad de datos dentro

#print(name)
#print(lename)

wb = openpyxl.Workbook()
ruta = 'Loterry.xlsx'

hoja = wb.active
hoja.title = "Name-Account-Country-Music"

fila = 1 #Fila donde empezamos
col_name = 1 #Columna donde guardamos el nombre
col_account = 2 #Columna donde guardamos el dinero
col_country = 3 #Columna de pais
col_music = 4 #Columna de musica

for name, dinero, pais in name.items():
    hoja.cell(column=col_name, row=fila, value=name)
    hoja.cell(column=col_account, row=fila, value=dato)
    hoja.cell(column=col_country, row=fila, value=dato)
    hoja.cell(column=col_music, row=fila, value=dato)
fila+=1

wb.save(filename = ruta)


